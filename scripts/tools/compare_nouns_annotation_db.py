import sqlite3
import logging
import time
import contextlib
import os
import pandas as pd
import click
import tqdm


logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

DISABLE_TIMEER = os.environ.get("DISABLE_TIMER", False)
DEBUG = os.environ.get("DEBUG", False)


@contextlib.contextmanager
def timer(timer_name="timer", pbar=None, pos=0):
    if DISABLE_TIMEER:
        return

    start = time.time()
    yield
    end = time.time()
    if pbar is not None:
        pbar.display(f"Time taken in [{timer_name}]: {end - start:.2f}", pos=pos)
    else:
        logger.info(f"Time taken in [{timer_name}]: {end - start:.2f}")


def get_tables_with_name_and_schema(cursor):
    # Get the list of tables
    cursor.execute("SELECT name, sql FROM sqlite_master WHERE type='table'")

    # Print the table names and their schema
    table_names = []
    table_schemas = []
    for result in cursor.fetchall():
        table_name, table_schema = result["name"], result["sql"]
        table_names.append(table_name)
        table_schemas.append(table_schema)

    return table_names, table_schemas


def load_rows(cursor, table_name):
    logger.info(f"Loading table: {table_name}")
    pos_table_name = table_name + "_pos_extension"
    cursor.execute(
        f"""  
        SELECT {table_name}.region_id, {table_name}.phrases, {pos_table_name}.nouns, {pos_table_name}.noun_chunks
        FROM {table_name}
        JOIN {pos_table_name} ON {table_name}.region_id = {pos_table_name}.region_id
    """
        + ("LIMIT 10" if DEBUG else "")
    )

    rows = cursor.fetchall()
    logger.info(f"Finished loading table: {table_name} with {len(rows)} rows")
    return rows


def dict_factory(cursor, row):
    # NOTE: now we will be returning rows as dictionaries instead of tuples
    d = {}
    for idx, col in enumerate(cursor.description):
        d[col[0]] = row[idx]
    return d


def load_db(db):
    if not os.path.exists(db):
        raise ValueError(f"db: {db} does not exist")
    conn = sqlite3.connect(db)

    conn.row_factory = dict_factory

    cursor = conn.cursor()
    table_names, _ = get_tables_with_name_and_schema(cursor)
    table_names = list(filter(lambda x: not x.endswith("_extension"), table_names))

    logger.info(f"Table Names: {table_names} from {db}")
    if len(table_names) != 2:
        raise ValueError(
            f"Expected 2 tables. The first one is for train, the second one is for eval, but we got {len(table_names)} tabls: {table_names}"
        )

    rows_ls = []
    noun2region_id_ls = []
    for table_name in table_names:
        with timer("load_rows"):
            rows = load_rows(cursor, table_name)
            rows_ls.append(rows)
        with timer("convert_rows_to_noun2region_id"):
            noun2region_id = convert_rows_to_noun2region_id(rows)
            noun2region_id_ls.append(noun2region_id)

    conn.close()

    for table_name, noun2region_id in zip(table_names, noun2region_id_ls):
        logger.info(f"# of nouns: {len(noun2region_id)} in {table_name}")
    return table_names, rows_ls, noun2region_id_ls


def convert_rows_to_noun2region_id(rows):
    noun2region_id = {}
    for row in rows:
        region_id = row["region_id"]
        phrases = row["phrases"]
        noun_chunks = row["noun_chunks"]
        nouns = row["nouns"]
        nouns = nouns.split("\t")
        for noun in nouns:
            if noun == "":
                # logger.warning(
                #     f"Empty noun in region_id: {region_id}, phrases: {phrases}, noun_chunks: {noun_chunks}, nouns: {nouns}"
                # )
                continue
            if noun not in noun2region_id:
                noun2region_id[noun] = []
            noun2region_id[noun].append(region_id)
    return noun2region_id


DB = "tmp/annotation_db/objects365-local/annotations.db"


@click.command()
@click.option("--db", help="Path to the database file", default=[DB], multiple=True)
@click.option(
    "--output_df_path",
    "-o",
    help="Path to the output dataframe",
    default="tmp/annotation_db_noun_stats/compare_nouns_annotation_db.xlsx",
)
def main(db, output_df_path):
    if len(db) != 2:
        raise ValueError(f"Expected 2 databases, but got {len(db)}: {db}")
    logger.info(f"Comparing db:")
    logger.info(f"\tA is: {db[0]}")
    logger.info(f"\tB is: {db[1]}")

    db_keys_ls = []
    for db_ in db:
        table_names, rows_ls, noun2region_id_ls = load_db(db_)
        db_keys = {
            table_name: noun2region_id.keys() for table_name, noun2region_id in zip(table_names, noun2region_id_ls)
        }
        assert len(noun2region_id_ls) == 2
        db_keys["+".join(table_names)] = noun2region_id_ls[0].keys() | noun2region_id_ls[1].keys()
        db_keys_ls.append(db_keys)

    for db_keys in db_keys_ls:
        assert len(db_keys) == 3

    ITEM_NAME = ["Train", "Eval", "Full"]
    compare_dict_ls = []
    for i in range(3):
        db_keys_a_dict = db_keys_ls[0]
        db_keys_a_names = list(db_keys_a_dict.keys())
        db_keys_a_name = db_keys_a_names[i]
        db_keys_a = db_keys_a_dict[db_keys_a_name]

        db_keys_b_dict = db_keys_ls[1]
        db_keys_b_names = list(db_keys_b_dict.keys())
        db_keys_b_name = db_keys_b_names[i]
        db_keys_b = db_keys_b_dict[db_keys_b_name]

        # prefix = ITEM_NAME[i] + "."
        prefix = ""
        compare_dict = {
            "name.A": db_keys_a_name,
            "name.B": db_keys_b_name,
            prefix + "A": len(db_keys_a),
            prefix + "B": len(db_keys_b),
            prefix + "A&B": len(db_keys_a & db_keys_b),
            prefix + "A|B": len(db_keys_a | db_keys_b),
            prefix + "A-B": len(db_keys_a - db_keys_b),
            prefix + "B-A": len(db_keys_b - db_keys_a),
        }
        compare_dict.update(
            {
                prefix + "A/A|B": compare_dict[prefix + "A"] / compare_dict[prefix + "A|B"],
                prefix + "B/A|B": compare_dict[prefix + "B"] / compare_dict[prefix + "A|B"],
                prefix + "A&B/A|B": compare_dict[prefix + "A&B"] / compare_dict[prefix + "A|B"],
            }
        )
        compare_dict_ls.append(compare_dict)

    df = [pd.DataFrame([compare_dict]) for compare_dict in compare_dict_ls]
    df = pd.concat(df, axis=1)
    os.makedirs(os.path.dirname(output_df_path), exist_ok=True)
    if not os.path.exists(output_df_path):
        logger.info(f"Saving dataframe to {output_df_path}")
        df.to_excel(output_df_path, index=False)
    else:
        from openpyxl import load_workbook

        with pd.ExcelWriter(output_df_path, engine="openpyxl", mode="a", if_sheet_exists="overlay") as writer:
            # Use workbook instead of book. https://stackoverflow.com/a/74449766
            writer.workbook = load_workbook(output_df_path)
            startrow = writer.sheets["Sheet1"].max_row
            df.to_excel(writer, startrow=startrow, index=False, header=False)
            logger.info(f"Append saving dataframe to {output_df_path}")


if __name__ == "__main__":
    main()
